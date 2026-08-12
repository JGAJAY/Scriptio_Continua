import math
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.cell import WriteOnlyCell  # type: ignore
from openpyxl.styles import Font  # type: ignore

try:
    from bert_score import BERTScorer  # type: ignore
except Exception:
    BERTScorer = None


BASE_DIR = Path(
    r"C:\Users\kamma\OneDrive - Amrita vishwa vidyapeetham\Project_Phase\DATASET\18000 rows\new dataset"
)
BASE_SENT_ID = BASE_DIR / "SENT_ID.xlsx"
X_VALUES = list(range(10, 101, 10))
LABEL_ORDER = ["B", "M", "E", "S"]
BERT_MODEL_TYPE = "roberta-large"
BERT_BATCH_SIZE = 16
_BERT_SCORER = None


def normalize_script(text: Optional[str]) -> str:
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch.lower() for ch in text if ch.isalnum())


def normalize_segmented_text(text: Optional[str]) -> str:
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^0-9A-Za-z\s]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().upper()


def parse_state4(value: Optional[str], n_chars: int) -> List[str]:
    tags = re.findall(r"[SBEMI]", str(value or "").upper())
    tags = ["M" if tag == "I" else tag for tag in tags]
    if len(tags) < n_chars:
        tags.extend(["M"] * (n_chars - len(tags)))
    return tags[:n_chars]


def segmented_to_bies(segmented_text: Optional[str], script_contin: str) -> List[str]:
    text = "" if segmented_text is None else str(segmented_text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    tokens = [normalize_script(token) for token in re.split(r"\s+", text.strip()) if token.strip()]
    tokens = [token for token in tokens if token]

    expected = normalize_script(script_contin)
    joined = "".join(tokens)

    if joined != expected:
        compact = normalize_script(segmented_text)
        if compact == expected:
            tokens = [compact] if compact else []
        else:
            tags = ["M"] * len(expected)
            if len(expected) == 1:
                tags[0] = "S"
                return tags
            if tags:
                tags[0] = "B"
                tags[-1] = "E"
            return tags

    tags: List[str] = []
    for token in tokens:
        if len(token) == 1:
            tags.append("S")
        else:
            tags.append("B")
            if len(token) > 2:
                tags.extend(["M"] * (len(token) - 2))
            tags.append("E")

    if len(tags) < len(expected):
        tags.extend(["M"] * (len(expected) - len(tags)))
    return tags[: len(expected)]


def lcs_length(a: Sequence[str], b: Sequence[str]) -> int:
    if not a or not b:
        return 0
    prev = [0] * (len(b) + 1)
    curr = [0] * (len(b) + 1)
    for i in range(1, len(a) + 1):
        for j in range(1, len(b) + 1):
            if a[i - 1] == b[j - 1]:
                curr[j] = prev[j - 1] + 1
            else:
                curr[j] = max(prev[j], curr[j - 1])
        prev, curr = curr, [0] * (len(b) + 1)
    return prev[len(b)]


def corpus_bleu_simple(references: Sequence[str], candidates: Sequence[str], max_n: int = 4) -> float:
    if not references:
        return 0.0
    eps = 1e-9
    clipped_counts = [0] * max_n
    total_counts = [0] * max_n
    ref_len = 0
    cand_len = 0

    for ref, cand in zip(references, candidates):
        ref_toks = ref.split()
        cand_toks = cand.split()
        ref_len += len(ref_toks)
        cand_len += len(cand_toks)
        for n in range(1, max_n + 1):
            ref_ngrams = Counter(tuple(ref_toks[i : i + n]) for i in range(max(0, len(ref_toks) - n + 1)))
            cand_ngrams = Counter(tuple(cand_toks[i : i + n]) for i in range(max(0, len(cand_toks) - n + 1)))
            total_counts[n - 1] += sum(cand_ngrams.values())
            clipped_counts[n - 1] += sum(min(count, ref_ngrams[gram]) for gram, count in cand_ngrams.items())

    precisions = [(clipped_counts[i] + eps) / (total_counts[i] + eps) for i in range(max_n)]
    if cand_len == 0:
        return 0.0
    bp = 1.0 if cand_len > ref_len else math.exp(1.0 - (ref_len / max(cand_len, 1)))
    return float(bp * math.exp(sum(math.log(p) for p in precisions) / max_n))


def rouge_l_avg(references: Sequence[str], candidates: Sequence[str]) -> float:
    scores: List[float] = []
    for ref, cand in zip(references, candidates):
        ref_toks = ref.split()
        cand_toks = cand.split()
        if not ref_toks and not cand_toks:
            scores.append(1.0)
            continue
        if not ref_toks or not cand_toks:
            scores.append(0.0)
            continue
        lcs = lcs_length(ref_toks, cand_toks)
        precision = lcs / len(cand_toks)
        recall = lcs / len(ref_toks)
        denom = precision + recall
        scores.append(0.0 if denom == 0 else (2 * precision * recall / denom))
    return sum(scores) / len(scores) if scores else 0.0


def meteor_like_avg(references: Sequence[str], candidates: Sequence[str]) -> float:
    scores: List[float] = []
    for ref, cand in zip(references, candidates):
        ref_toks = ref.split()
        cand_toks = cand.split()
        if not ref_toks and not cand_toks:
            scores.append(1.0)
            continue
        if not ref_toks or not cand_toks:
            scores.append(0.0)
            continue
        ref_counts = Counter(ref_toks)
        cand_counts = Counter(cand_toks)
        matches = sum(min(ref_counts[word], cand_counts[word]) for word in cand_counts)
        precision = matches / len(cand_toks)
        recall = matches / len(ref_toks)
        denom = recall + 9 * precision
        scores.append(0.0 if denom == 0 else (10 * precision * recall / denom))
    return sum(scores) / len(scores) if scores else 0.0


def classification_metrics(y_true: Sequence[str], y_pred: Sequence[str]) -> Dict[str, float]:
    total = len(y_true)
    if total == 0:
        return {"precision": 0.0, "recall": 0.0, "f1": 0.0, "accuracy": 0.0}

    support = Counter(y_true)
    predicted = Counter(y_pred)
    true_positive = Counter(label for label, pred in zip(y_true, y_pred) if label == pred)
    weighted_precision = 0.0
    weighted_recall = 0.0
    weighted_f1 = 0.0

    for label in LABEL_ORDER:
        tp = true_positive[label]
        fp = predicted[label] - tp
        fn = support[label] - tp
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0
        weighted_precision += precision * support[label]
        weighted_recall += recall * support[label]
        weighted_f1 += f1 * support[label]

    accuracy = sum(1 for truth, pred in zip(y_true, y_pred) if truth == pred) / total
    return {
        "precision": weighted_precision / total,
        "recall": weighted_recall / total,
        "f1": weighted_f1 / total,
        "accuracy": accuracy,
    }


def text_metrics_for_pair(reference: str, candidate: str) -> Dict[str, float]:
    return {
        "BLEU": corpus_bleu_simple([reference], [candidate]),
        "ROUGE": rouge_l_avg([reference], [candidate]),
        "METEOR": meteor_like_avg([reference], [candidate]),
    }


def populate_bert_cache(bert_cache: Dict[Tuple[str, str], object], pairs: Sequence[Tuple[str, str]]) -> None:
    missing_pairs = [pair for pair in pairs if pair not in bert_cache]
    if not missing_pairs:
        return

    global _BERT_SCORER
    if BERTScorer is None:
        for pair in missing_pairs:
            bert_cache[pair] = "N/A (bert-score import failed)"
        return

    references = [pair[0] for pair in missing_pairs]
    candidates = [pair[1] for pair in missing_pairs]
    try:
        if _BERT_SCORER is None:
            _BERT_SCORER = BERTScorer(lang="en", model_type=BERT_MODEL_TYPE, batch_size=BERT_BATCH_SIZE)
        _, _, f1 = _BERT_SCORER.score(candidates, references)
        for pair, value in zip(missing_pairs, f1.tolist()):
            bert_cache[pair] = float(value)
    except Exception as exc:
        message = f"N/A ({type(exc).__name__}: {exc})"
        for pair in missing_pairs:
            bert_cache[pair] = message


def build_base_lookup() -> Dict[Tuple[str, str, str, str], Dict[str, str]]:
    lookup: Dict[Tuple[str, str, str, str], Dict[str, str]] = {}
    wb = load_workbook(BASE_SENT_ID, read_only=True, data_only=True)
    try:
        for split_name in wb.sheetnames:
            ws = wb[split_name]
            header = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            index = {name: idx for idx, name in enumerate(header)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                art_id = str(row[index["ART_ID"]] or "")
                para_id = str(row[index["PARA_ID"]] or "")
                sent_id = str(row[index["SENT_ID"]] or "")
                lookup[(split_name, art_id, para_id, sent_id)] = {
                    "script_contin": str(row[index["SCRIPT_CONTIN"]] or ""),
                    "state_4": str(row[index["STATE_4"]] or ""),
                }
    finally:
        wb.close()
    return lookup


def iter_output_rows(ws) -> Iterable[Dict[str, object]]:
    header = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: idx for idx, name in enumerate(header)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        output = row[index["output"]] if "output" in index else None
        if output is None or str(output).strip() == "":
            continue
        yield {
            "split_name": str(row[index["sheet_name"]] or ""),
            "art_id": str(row[index["ART_ID"]] or ""),
            "para_id": str(row[index["PARA_ID"]] or ""),
            "sent_id": str(row[index["SENT_ID"]] or ""),
            "ground_truth": str(row[index["ground_truth"]] or ""),
            "output": str(output or ""),
        }


def build_metric_rows(
    result_file: Path,
    metric_label: str,
    base_lookup: Dict[Tuple[str, str, str, str], Dict[str, str]],
    output_file: Path,
) -> None:
    source_wb = load_workbook(result_file, read_only=True, data_only=True)
    out_wb = Workbook(write_only=True)
    try:
        for x in X_VALUES:
            rows_for_sheet: List[List[object]] = []
            references: List[str] = []
            candidates: List[str] = []
            y_true: List[str] = []
            y_pred: List[str] = []

            for split_name in ["train", "test", "val"]:
                source_sheet_name = f"{split_name}_x{x}"
                if source_sheet_name not in source_wb.sheetnames:
                    continue
                ws = source_wb[source_sheet_name]
                for result_row in iter_output_rows(ws):
                    key = (
                        str(result_row["split_name"]),
                        str(result_row["art_id"]),
                        str(result_row["para_id"]),
                        str(result_row["sent_id"]),
                    )
                    base_row = base_lookup.get(key)
                    if base_row is None:
                        continue

                    script_contin = base_row["script_contin"]
                    gt_bies = parse_state4(base_row["state_4"], len(script_contin))
                    out_bies = segmented_to_bies(str(result_row["output"]), script_contin)

                    ground_truth = normalize_segmented_text(result_row["ground_truth"])
                    output = normalize_segmented_text(result_row["output"])
                    references.append(ground_truth)
                    candidates.append(output)
                    y_true.extend(gt_bies)
                    y_pred.extend(out_bies)

                    rows_for_sheet.append(
                        [
                            str(result_row["sent_id"]),
                            script_contin,
                            ground_truth,
                            output,
                            " ".join(gt_bies),
                            " ".join(out_bies),
                        ]
                    )

            ws_out = out_wb.create_sheet(title=f"x{x}")
            header = [
                "sent id",
                "scripto continua(input sentence)",
                "ground truth",
                "output sentence",
                "ground truth 4 state (BIES)",
                "output 4 state (BIES)",
            ]
            header_cells = []
            for value in header:
                cell = WriteOnlyCell(ws_out, value=value)
                cell.font = Font(bold=True)
                header_cells.append(cell)
            ws_out.append(header_cells)

            for row in rows_for_sheet:
                ws_out.append(row)

            class_scores = classification_metrics(y_true, y_pred)
            metric_rows = [
                [],
                ["metric", "value"],
                ["BLEU", corpus_bleu_simple(references, candidates)],
                ["ROUGE", rouge_l_avg(references, candidates)],
                ["METEOR", meteor_like_avg(references, candidates)],
                ["BERT", "N/A (bert-score not available in current environment)"],
                ["F1", class_scores["f1"]],
                ["PRECISION", class_scores["precision"]],
                ["ACCURACY", class_scores["accuracy"]],
                ["RECALL", class_scores["recall"]],
                ["SOURCE_METRIC_COLUMN", metric_label],
                ["ROW_COUNT", len(rows_for_sheet)],
            ]
            for idx, row in enumerate(metric_rows):
                if idx == 1:
                    cells = []
                    for value in row:
                        cell = WriteOnlyCell(ws_out, value=value)
                        cell.font = Font(bold=True)
                        cells.append(cell)
                    ws_out.append(cells)
                else:
                    ws_out.append(row)

        out_wb.save(output_file)
    finally:
        source_wb.close()


def build_sentence_metric_rows(
    result_file: Path,
    metric_label: str,
    base_lookup: Dict[Tuple[str, str, str, str], Dict[str, str]],
    output_file: Path,
    split_names: Optional[Sequence[str]] = None,
) -> None:
    source_wb = load_workbook(result_file, read_only=True, data_only=True)
    out_wb = Workbook(write_only=True)
    bert_cache: Dict[Tuple[str, str], object] = {}
    active_splits = list(split_names or ["train", "test", "val"])
    try:
        for x in X_VALUES:
            ws_out = out_wb.create_sheet(title=f"x{x}")
            header = [
                "sent id",
                "scripto continua(input sentence)",
                "ground truth",
                "output sentence",
                "ground truth 4 state (BIES)",
                "output 4 state (BIES)",
                "BLEU",
                "ROUGE",
                "METEOR",
                "BERT",
                "F1",
                "PRECISION",
                "ACCURACY",
                "RECALL",
                "SOURCE_METRIC_COLUMN",
            ]
            header_cells = []
            for value in header:
                cell = WriteOnlyCell(ws_out, value=value)
                cell.font = Font(bold=True)
                header_cells.append(cell)
            ws_out.append(header_cells)

            sheet_rows: List[List[object]] = []
            bert_pairs: List[Tuple[str, str]] = []
            for split_name in active_splits:
                source_sheet_name = f"{split_name}_x{x}"
                if source_sheet_name not in source_wb.sheetnames:
                    continue
                ws = source_wb[source_sheet_name]
                for result_row in iter_output_rows(ws):
                    key = (
                        str(result_row["split_name"]),
                        str(result_row["art_id"]),
                        str(result_row["para_id"]),
                        str(result_row["sent_id"]),
                    )
                    base_row = base_lookup.get(key)
                    if base_row is None:
                        continue

                    script_contin = base_row["script_contin"]
                    gt_bies = parse_state4(base_row["state_4"], len(script_contin))
                    out_bies = segmented_to_bies(str(result_row["output"]), script_contin)

                    ground_truth = normalize_segmented_text(result_row["ground_truth"])
                    output = normalize_segmented_text(result_row["output"])
                    text_scores = text_metrics_for_pair(ground_truth, output)
                    class_scores = classification_metrics(gt_bies, out_bies)
                    bert_pair = (ground_truth, output)
                    bert_pairs.append(bert_pair)
                    sheet_rows.append(
                        [
                            str(result_row["sent_id"]),
                            script_contin,
                            ground_truth,
                            output,
                            " ".join(gt_bies),
                            " ".join(out_bies),
                            text_scores["BLEU"],
                            text_scores["ROUGE"],
                            text_scores["METEOR"],
                            bert_pair,
                            class_scores["f1"],
                            class_scores["precision"],
                            class_scores["accuracy"],
                            class_scores["recall"],
                            metric_label,
                        ]
                    )

            populate_bert_cache(bert_cache, bert_pairs)
            for row in sheet_rows:
                bert_pair = row[9]
                row[9] = bert_cache.get(bert_pair, "N/A (bert-score unavailable)")
                ws_out.append(row)

        out_wb.save(output_file)
    finally:
        source_wb.close()


def main() -> None:
    base_lookup = build_base_lookup()
    jobs = [
        ("SENT_ID_coh_results.xlsx", "coh", "coh_metrics.xlsx"),
        ("SENT_ID_gs_results.xlsx", "gs", "gs_metrics.xlsx"),
        ("SENT_ID_lm_results.xlsx", "lm", "lm_metrics.xlsx"),
    ]
    for source_name, metric_label, output_name in jobs:
        build_metric_rows(
            result_file=BASE_DIR / source_name,
            metric_label=metric_label,
            base_lookup=base_lookup,
            output_file=BASE_DIR / output_name,
        )
        print(f"Created {output_name}")

    sentence_jobs = [
        ("SENT_ID_coh_results.xlsx", "coh", "coh1_metrics.xlsx"),
        ("SENT_ID_gs_results.xlsx", "gs", "gs1_metrics.xlsx"),
        ("SENT_ID_lm_results.xlsx", "lm", "lm1_metrics.xlsx"),
    ]
    for source_name, metric_label, output_name in sentence_jobs:
        build_sentence_metric_rows(
            result_file=BASE_DIR / source_name,
            metric_label=metric_label,
            base_lookup=base_lookup,
            output_file=BASE_DIR / output_name,
        )
        print(f"Created {output_name}")

    test_only_sentence_jobs = [
        ("SENT_ID_coh_results.xlsx", "coh", "coh2_metrics.xlsx"),
        ("SENT_ID_gs_results.xlsx", "gs", "gs2_metrics.xlsx"),
        ("SENT_ID_lm_results.xlsx", "lm", "lm2_metrics.xlsx"),
    ]
    for source_name, metric_label, output_name in test_only_sentence_jobs:
        build_sentence_metric_rows(
            result_file=BASE_DIR / source_name,
            metric_label=metric_label,
            base_lookup=base_lookup,
            output_file=BASE_DIR / output_name,
            split_names=["test"],
        )
        print(f"Created {output_name}")


if __name__ == "__main__":
    main()
