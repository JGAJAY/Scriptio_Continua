from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


BASE_DIR = Path(
    r"C:\Users\kamma\OneDrive - Amrita vishwa vidyapeetham\Project_Phase\DATASET\18000 rows\new dataset"
)

FILES = [
    BASE_DIR / "coh2_metrics.xlsx",
    BASE_DIR / "gs2_metrics.xlsx",
    BASE_DIR / "lm2_metrics.xlsx",
]

METRIC_COLUMNS = {
    "BLEU",
    "ROUGE",
    "METEOR",
    "BERT",
    "F1",
    "PRECISION",
    "ACCURACY",
    "RECALL",
}
SUMMARY_ORDER = ["BLEU", "ROUGE", "METEOR", "BERT", "F1", "PRECISION", "ACCURACY", "RECALL"]


def is_number(value) -> bool:
    return isinstance(value, (int, float))


def process_file(path: Path) -> None:
    wb = load_workbook(path)
    summary_rows = []
    for ws in wb.worksheets:
        if ws.title.lower() == "summary":
            continue
        header = [cell.value for cell in ws[1]]
        header_index = {str(value): idx + 1 for idx, value in enumerate(header) if value is not None}

        if ws.max_row >= 2 and str(ws.cell(ws.max_row, 1).value).strip().upper() == "AVG":
            ws.delete_rows(ws.max_row, 1)

        avg_row = [""] * len(header)
        avg_row[0] = "AVG"

        for col_name in METRIC_COLUMNS:
            col_idx = header_index.get(col_name)
            if not col_idx:
                continue
            values = []
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row_idx, col_idx).value
                if is_number(value):
                    values.append(float(value))
            avg_row[col_idx - 1] = (sum(values) / len(values)) if values else None

        ws.append(avg_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        summary_rows.append(
            [ws.title]
            + [avg_row[header_index[col_name] - 1] if col_name in header_index else None for col_name in SUMMARY_ORDER]
        )

    if "summary" in wb.sheetnames:
        del wb["summary"]

    summary_ws = wb.create_sheet("summary")
    summary_header = ["sheet name"] + SUMMARY_ORDER
    summary_ws.append(summary_header)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)

    def sort_key(row):
        name = str(row[0])
        if name.startswith("x") and name[1:].isdigit():
            return int(name[1:])
        return 10**9

    for row in sorted(summary_rows, key=sort_key):
        summary_ws.append(row)

    wb.save(path)


def main() -> None:
    for path in FILES:
        process_file(path)
        print(f"Updated {path.name}")


if __name__ == "__main__":
    main()
