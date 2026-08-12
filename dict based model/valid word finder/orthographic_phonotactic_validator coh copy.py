"""
Orthographic phonotactic validator plus CM-only and COH-only sentence experiment export.

How this file works:
1. Train a word-level orthographic validator from weighted words.
2. Optionally score a single word or evaluate a word dataset.
3. Optionally run a CM-only or COH-only sentence reconstruction experiment over
   `SCRIPT_CONTIN` sentences, testing x = 10, 20, ..., 100.

The CM mode uses the context model's built-in scoring.
The COH mode uses PMI-based cohesion scoring over all generated candidates.
"""

from __future__ import annotations

import argparse
import math
import re
import statistics
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from wordfreq import iter_wordlist, zipf_frequency

SIBLING_MODEL_DIR = Path(__file__).resolve().parent / "word prob finding models"
if str(SIBLING_MODEL_DIR) not in sys.path:
    sys.path.insert(0, str(SIBLING_MODEL_DIR))

from character_backoff_pronounceability_validator import VOWELS, split_into_sound_chunks
from character_backoff_language_model import load_words, make_readable_copy, normalize_word


STATE4_RE = re.compile(r"[SBEMI]")


def logistic(value: float) -> float:
    if value >= 0:
        return 1.0 / (1.0 + math.exp(-value))
    exp_value = math.exp(value)
    return exp_value / (1.0 + exp_value)


def zscore_to_unit(value: float, mean: float, std: float) -> float:
    return logistic((value - mean) / std)


class PhonotacticWordlikenessValidator:
    def __init__(self, max_ngram: int = 4, alpha: float = 0.1) -> None:
        self.max_ngram = max_ngram
        self.alpha = alpha
        self.ngram_counts: dict[int, Counter[str]] = {n: Counter() for n in range(1, max_ngram + 1)}
        self.ngram_totals: dict[int, int] = {}
        self.chunk_counts: Counter[str] = Counter()
        self.vocab_letters: set[str] = set()
        self.familiarity_lexicon: dict[str, float] = {}
        self.train_ngram_mean = -8.0
        self.train_ngram_std = 1.0
        self.train_chunk_mean = -6.0
        self.train_chunk_std = 1.0

    def fit(self, weighted_words: list[tuple[str, int]]) -> None:
        self.ngram_counts = {n: Counter() for n in range(1, self.max_ngram + 1)}
        self.chunk_counts = Counter()
        self.vocab_letters = set()

        for word, count in weighted_words:
            self.vocab_letters.update(word)
            padded = f"^{word}$"
            for n in range(1, self.max_ngram + 1):
                if len(padded) < n:
                    continue
                for i in range(len(padded) - n + 1):
                    self.ngram_counts[n][padded[i:i + n]] += count

            for n in range(2, 5):
                if len(word) < n:
                    continue
                for i in range(len(word) - n + 1):
                    self.chunk_counts[word[i:i + n]] += count

        self.ngram_totals = {n: sum(counter.values()) for n, counter in self.ngram_counts.items()}
        self.familiarity_lexicon = self._build_wordfreq_lexicon()
        self._fit_score_calibration(weighted_words)

    def _build_wordfreq_lexicon(self) -> dict[str, float]:
        lexicon: dict[str, float] = {}
        for raw_word in iter_wordlist("en", wordlist="large"):
            word = normalize_word(raw_word)
            if not word:
                continue
            lexicon[word] = zipf_frequency(word, "en", wordlist="large")
        return lexicon

    def _smoothed_ngram_prob(self, gram: str) -> float:
        n = len(gram)
        total = self.ngram_totals.get(n, 0)
        count = self.ngram_counts[n].get(gram, 0)
        vocab_size = max(1, len(self.vocab_letters) + 2)
        return (count + self.alpha) / (total + self.alpha * (vocab_size ** n))

    def _word_ngram_logscore(self, word: str) -> float:
        padded = f"^{word}$"
        log_sum = 0.0
        grams = 0
        for n in range(2, self.max_ngram + 1):
            if len(padded) < n:
                continue
            weight = n - 1
            for i in range(len(padded) - n + 1):
                prob = self._smoothed_ngram_prob(padded[i:i + n])
                log_sum += weight * math.log(prob)
                grams += weight
        return log_sum / max(1, grams)

    def _word_chunk_logscore(self, word: str) -> float:
        values: list[float] = []
        for n in range(2, 5):
            if len(word) < n:
                continue
            for i in range(len(word) - n + 1):
                chunk = word[i:i + n]
                count = self.chunk_counts.get(chunk, 0)
                values.append(math.log(count + 1.0))
        if not values:
            return 0.0
        return sum(values) / len(values)

    def _shape_score(self, word: str) -> float:
        chunks = split_into_sound_chunks(word)
        if not chunks:
            return 0.0
        if not any(char in VOWELS for char in word):
            return 0.0

        consonant_runs = [len(chunk) for kind, chunk in chunks if kind == "C"]
        vowel_runs = [len(chunk) for kind, chunk in chunks if kind == "V"]

        score = 1.0
        if consonant_runs:
            score -= 0.18 * max(0, max(consonant_runs) - 3)
        if vowel_runs:
            score -= 0.15 * max(0, max(vowel_runs) - 2)
        if len(chunks) == 1:
            score -= 0.25
        if word.endswith("q") or word.startswith("q") and len(word) == 1:
            score -= 0.2
        return max(0.0, min(1.0, score))

    def _familiarity_score(self, word: str) -> float:
        zipf = self.familiarity_lexicon.get(word, 0.0)
        return max(0.0, min(1.0, zipf / 8.0))

    def _fit_score_calibration(self, weighted_words: list[tuple[str, int]]) -> None:
        ngram_values: list[float] = []
        chunk_values: list[float] = []
        for word, count in weighted_words:
            ngram_score = self._word_ngram_logscore(word)
            chunk_score = self._word_chunk_logscore(word)
            repeats = min(count, 20)
            ngram_values.extend([ngram_score] * repeats)
            chunk_values.extend([chunk_score] * repeats)

        if ngram_values:
            self.train_ngram_mean = statistics.mean(ngram_values)
            self.train_ngram_std = statistics.pstdev(ngram_values) or 1.0
        if chunk_values:
            self.train_chunk_mean = statistics.mean(chunk_values)
            self.train_chunk_std = statistics.pstdev(chunk_values) or 1.0

    def score_word(self, raw_word: str) -> dict[str, float | bool | str]:
        word = normalize_word(raw_word)
        if not word:
            return {
                "word": str(raw_word),
                "score": 0.0,
                "is_wordlike": False,
                "ngram_score": 0.0,
                "chunk_score": 0.0,
                "shape_score": 0.0,
                "familiarity_score": 0.0,
            }

        raw_ngram = self._word_ngram_logscore(word)
        raw_chunk = self._word_chunk_logscore(word)
        ngram_score = zscore_to_unit(raw_ngram, self.train_ngram_mean, self.train_ngram_std)
        chunk_score = zscore_to_unit(raw_chunk, self.train_chunk_mean, self.train_chunk_std)
        shape_score = self._shape_score(word)
        familiarity_score = self._familiarity_score(word)

        combined = (
            0.45 * ngram_score
            + 0.25 * chunk_score
            + 0.15 * shape_score
            + 0.15 * familiarity_score
        )
        return {
            "word": word,
            "score": combined,
            "is_wordlike": combined >= 0.5,
            "ngram_score": ngram_score,
            "chunk_score": chunk_score,
            "shape_score": shape_score,
            "familiarity_score": familiarity_score,
        }


def evaluate_dataset(path: Path, validator: PhonotacticWordlikenessValidator) -> dict[str, float | int]:
    words = load_words(path)
    total_rows = 0
    total_weight = 0
    score_sum = 0.0
    weighted_score_sum = 0.0
    accepted_rows = 0
    accepted_weight = 0

    for word, count in words:
        result = validator.score_word(word)
        score = float(result["score"])
        total_rows += 1
        total_weight += count
        score_sum += score
        weighted_score_sum += score * count
        if result["is_wordlike"]:
            accepted_rows += 1
            accepted_weight += count

    return {
        "rows": total_rows,
        "total_weight": total_weight,
        "accepted_row_fraction": accepted_rows / total_rows if total_rows else 0.0,
        "accepted_weight_fraction": accepted_weight / total_weight if total_weight else 0.0,
        "average_score": score_sum / total_rows if total_rows else 0.0,
        "weighted_average_score": weighted_score_sum / total_weight if total_weight else 0.0,
    }


def normalize_script_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    return "".join(ch for ch in text if ch.isalpha() or ch.isdigit())


def parse_state4(value: object, n_chars: int) -> list[str]:
    tags = STATE4_RE.findall(str(value).upper())
    tags = ["M" if tag == "I" else tag for tag in tags]
    if len(tags) < n_chars:
        tags += ["M"] * (n_chars - len(tags))
    return tags[:n_chars]


def tokens_from_state4(script: str, raw_labels: object) -> list[str]:
    labels = parse_state4(raw_labels, len(script))
    tokens: list[str] = []
    current: list[str] = []

    for index, char in enumerate(script):
        tag = labels[index]
        if tag == "S":
            if current:
                tokens.append("".join(current))
                current = []
            tokens.append(char)
        elif tag == "B":
            if current:
                tokens.append("".join(current))
            current = [char]
        elif tag == "E":
            if current:
                current.append(char)
                tokens.append("".join(current))
                current = []
            else:
                tokens.append(char)
        else:
            if current:
                current.append(char)
            else:
                current = [char]

    if current:
        tokens.append("".join(current))
    return [token for token in tokens if token]


@dataclass(frozen=True)
class CandidateArc:
    token: str
    end: int
    token_score: float
    is_valid: bool


@dataclass
class BeamState:
    position: int
    tokens: tuple[str, ...]
    word_score_sum: float
    valid_count: int
    bigram_raw_sum: float
    bigram_steps: int
    approx_score: float


class SentenceContextModel:
    def __init__(self, validator: PhonotacticWordlikenessValidator, alpha: float = 0.5) -> None:
        self.validator = validator
        self.alpha = alpha
        self.token_counts: Counter[str] = Counter()
        self.start_counts: Counter[str] = Counter()
        self.bigram_counts: Counter[tuple[str, str]] = Counter()
        self.prev_totals: Counter[str] = Counter()
        self.num_sentences = 0
        self.avg_token_length_mean = 4.5
        self.avg_token_length_std = 1.0
        self.bigram_raw_mean = -6.0
        self.bigram_raw_std = 1.0
        self.total_tokens = 0
        self.total_bigrams = 0
        self.vocab_size = 0

    @lru_cache(maxsize=200_000)
    def word_profile(self, token: str) -> tuple[float, bool]:
        if not token:
            return 0.0, False
        if token.isdigit():
            return 1.0, True

        result = self.validator.score_word(token)
        score = float(result["score"])
        is_valid = bool(result["is_wordlike"])

        # Keep single letters as usable initials instead of discarding them.
        if len(token) == 1 and token.isalpha():
            score = max(score, 0.55)
            is_valid = True

        return score, is_valid

    def fit(self, train_token_sequences: list[list[str]]) -> None:
        self.token_counts = Counter()
        self.start_counts = Counter()
        self.bigram_counts = Counter()
        self.prev_totals = Counter()
        self.num_sentences = len(train_token_sequences)

        avg_token_lengths: list[float] = []
        bigram_raw_values: list[float] = []

        for tokens in train_token_sequences:
            if not tokens:
                continue

            self.start_counts[tokens[0]] += 1
            self.token_counts.update(tokens)

            for prev, nxt in zip(tokens, tokens[1:]):
                self.bigram_counts[(prev, nxt)] += 1
                self.prev_totals[prev] += 1

        self.total_tokens = sum(self.token_counts.values())
        self.total_bigrams = sum(self.bigram_counts.values())
        self.vocab_size = len(self.token_counts)

        for tokens in train_token_sequences:
            if not tokens:
                continue
            avg_token_lengths.append(sum(len(token) for token in tokens) / len(tokens))
            bigram_raw_values.append(self._average_bigram_raw(tokens))

        if avg_token_lengths:
            self.avg_token_length_mean = statistics.mean(avg_token_lengths)
            self.avg_token_length_std = statistics.pstdev(avg_token_lengths) or 1.0
        if bigram_raw_values:
            self.bigram_raw_mean = statistics.mean(bigram_raw_values)
            self.bigram_raw_std = statistics.pstdev(bigram_raw_values) or 1.0

    def _start_raw(self, token: str) -> float:
        vocab_size = max(1, len(self.token_counts))
        total = max(1, self.num_sentences)
        count = self.start_counts.get(token, 0)
        probability = (count + self.alpha) / (total + self.alpha * vocab_size)
        return math.log(probability)

    def transition_raw(self, previous: str | None, token: str) -> float:
        if previous is None:
            return self._start_raw(token)

        vocab_size = max(1, len(self.token_counts))
        total = self.prev_totals.get(previous, 0)
        count = self.bigram_counts.get((previous, token), 0)
        probability = (count + self.alpha) / (total + self.alpha * vocab_size)
        return math.log(probability)

    def _average_bigram_raw(self, tokens: list[str] | tuple[str, ...]) -> float:
        if not tokens:
            return self.bigram_raw_mean
        values: list[float] = []
        previous: str | None = None
        for token in tokens:
            values.append(self.transition_raw(previous, token))
            previous = token
        return sum(values) / len(values)

    def compute_pmi(self, x: str, y: str) -> float:
        alpha = 0.5
        n = self.total_tokens
        m = self.total_bigrams
        v = self.vocab_size

        px = (self.token_counts[x] + alpha) / (n + alpha * v)
        py = (self.token_counts[y] + alpha) / (n + alpha * v)
        pxy = (self.bigram_counts.get((x, y), 0) + alpha) / (m + alpha * v)

        return math.log(pxy / (px * py))

    def compute_coh_score(self, tokens: list[str] | tuple[str, ...]) -> float:
        if len(tokens) < 2:
            return 0.0
        score = 0.0
        for i in range(len(tokens) - 1):
            score += self.compute_pmi(tokens[i], tokens[i + 1])
        return score

    def score_sentence(self, tokens: list[str] | tuple[str, ...]) -> dict[str, float | int | str]:
        token_list = list(tokens)
        if not token_list:
            return {
                "output_sentence": "",
                "cm_score": 0.0,
                "word_validity": 0.0,
                "bigram_validity": 0.0,
                "character_smoothness": 0.0,
                "sentence_length_sanity": 0.0,
                "mean_ortho_score": 0.0,
                "token_count": 0,
            }

        score_sum = 0.0
        valid_count = 0
        for token in token_list:
            score, is_valid = self.word_profile(token)
            score_sum += score
            valid_count += int(is_valid)

        word_validity = valid_count / len(token_list)
        character_smoothness = score_sum / len(token_list)
        bigram_raw = self._average_bigram_raw(token_list)
        bigram_validity = zscore_to_unit(bigram_raw, self.bigram_raw_mean, self.bigram_raw_std)
        avg_token_length = sum(len(token) for token in token_list) / len(token_list)
        sentence_length_sanity = zscore_to_unit(
            avg_token_length,
            self.avg_token_length_mean,
            self.avg_token_length_std,
        )

        cm_score = (
            0.35 * word_validity
            + 0.25 * bigram_validity
            + 0.25 * character_smoothness
            + 0.15 * sentence_length_sanity
        )

        return {
            "output_sentence": " ".join(token_list),
            "cm_score": cm_score,
            "word_validity": word_validity,
            "bigram_validity": bigram_validity,
            "character_smoothness": character_smoothness,
            "sentence_length_sanity": sentence_length_sanity,
            "mean_ortho_score": character_smoothness,
            "token_count": len(token_list),
        }


def build_candidate_arcs(
    script: str,
    context_model: SentenceContextModel,
    max_word_len: int,
) -> dict[int, list[CandidateArc]]:
    candidates: dict[int, list[CandidateArc]] = {}
    n_chars = len(script)
    index = 0

    while index < n_chars:
        char = script[index]
        if char.isdigit():
            end = index + 1
            while end < n_chars and script[end].isdigit():
                end += 1
            candidates[index] = [CandidateArc(token=script[index:end], end=end, token_score=1.0, is_valid=True)]
            for inner in range(index + 1, end):
                candidates[inner] = []
            index = end
            continue

        arcs: list[CandidateArc] = []
        limit = min(n_chars, index + max_word_len)
        for end in range(index + 1, limit + 1):
            token = script[index:end]
            if not token.isalpha():
                break
            score, is_valid = context_model.word_profile(token)
            arcs.append(CandidateArc(token=token, end=end, token_score=score, is_valid=is_valid))

        arcs.sort(key=lambda item: (-item.token_score, len(item.token), item.token))
        candidates[index] = arcs
        index += 1

    return candidates


def _approximate_state_score(
    state: BeamState,
    context_model: SentenceContextModel,
) -> float:
    if not state.tokens:
        return 0.0

    word_validity = state.valid_count / len(state.tokens)
    character_smoothness = state.word_score_sum / len(state.tokens)
    bigram_raw = state.bigram_raw_sum / max(1, state.bigram_steps)
    bigram_validity = zscore_to_unit(bigram_raw, context_model.bigram_raw_mean, context_model.bigram_raw_std)
    avg_token_length = state.position / len(state.tokens)
    length_sanity = zscore_to_unit(
        avg_token_length,
        context_model.avg_token_length_mean,
        context_model.avg_token_length_std,
    )
    return (
        0.35 * word_validity
        + 0.25 * bigram_validity
        + 0.25 * character_smoothness
        + 0.15 * length_sanity
    )


def search_best_sentence(
    script: str,
    candidates: dict[int, list[CandidateArc]],
    context_model: SentenceContextModel,
    top_k: int,
    beam_cap: int,
) -> tuple[dict[str, float | int | str], int]:
    n_chars = len(script)
    beam_width = max(32, min(beam_cap, top_k * 8))

    beam = [
        BeamState(
            position=0,
            tokens=(),
            word_score_sum=0.0,
            valid_count=0,
            bigram_raw_sum=0.0,
            bigram_steps=0,
            approx_score=0.0,
        )
    ]
    completed: list[BeamState] = []

    while beam:
        next_beam: list[BeamState] = []
        all_done = True

        for state in beam:
            if state.position >= n_chars:
                completed.append(state)
                continue

            all_done = False
            arcs = candidates.get(state.position, [])
            if not arcs:
                continue

            previous = state.tokens[-1] if state.tokens else None
            for arc in arcs[:top_k]:
                bigram_raw = context_model.transition_raw(previous, arc.token)
                next_state = BeamState(
                    position=arc.end,
                    tokens=state.tokens + (arc.token,),
                    word_score_sum=state.word_score_sum + arc.token_score,
                    valid_count=state.valid_count + int(arc.is_valid),
                    bigram_raw_sum=state.bigram_raw_sum + bigram_raw,
                    bigram_steps=state.bigram_steps + 1,
                    approx_score=0.0,
                )
                next_state.approx_score = _approximate_state_score(next_state, context_model)
                next_beam.append(next_state)

        if all_done:
            break
        if not next_beam:
            break

        next_beam.sort(key=lambda state: state.approx_score, reverse=True)
        beam = next_beam[:beam_width]

    final_states = completed or [state for state in beam if state.position >= n_chars]
    if not final_states:
        empty = context_model.score_sentence(())
        empty["output_sentence"] = script
        return empty, 0

    scored_states = []
    for state in final_states:
        metrics = context_model.score_sentence(state.tokens)
        scored_states.append((float(metrics["cm_score"]), metrics))
    scored_states.sort(key=lambda item: item[0], reverse=True)
    best_metrics = scored_states[0][1]
    return best_metrics, len(final_states)


def search_all_candidates(
    script: str,
    candidates: dict[int, list[CandidateArc]],
    context_model: SentenceContextModel,
    top_k: int,
    beam_cap: int,
) -> list[tuple[str, ...]]:
    n_chars = len(script)
    beam_width = max(32, min(beam_cap, top_k * 8))

    beam = [
        BeamState(
            position=0,
            tokens=(),
            word_score_sum=0.0,
            valid_count=0,
            bigram_raw_sum=0.0,
            bigram_steps=0,
            approx_score=0.0,
        )
    ]
    completed: list[BeamState] = []

    while beam:
        next_beam: list[BeamState] = []
        all_done = True

        for state in beam:
            if state.position >= n_chars:
                completed.append(state)
                continue

            all_done = False
            arcs = candidates.get(state.position, [])
            if not arcs:
                continue

            previous = state.tokens[-1] if state.tokens else None
            for arc in arcs[:top_k]:
                bigram_raw = context_model.transition_raw(previous, arc.token)
                next_state = BeamState(
                    position=arc.end,
                    tokens=state.tokens + (arc.token,),
                    word_score_sum=state.word_score_sum + arc.token_score,
                    valid_count=state.valid_count + int(arc.is_valid),
                    bigram_raw_sum=state.bigram_raw_sum + bigram_raw,
                    bigram_steps=state.bigram_steps + 1,
                    approx_score=0.0,
                )
                next_state.approx_score = _approximate_state_score(next_state, context_model)
                next_beam.append(next_state)

        if all_done:
            break
        if not next_beam:
            break

        next_beam.sort(key=lambda state: state.approx_score, reverse=True)
        beam = next_beam[:beam_width]

    final_states = completed or [state for state in beam if state.position >= n_chars]
    unique_token_sequences = list(dict.fromkeys(state.tokens for state in final_states))
    return unique_token_sequences


def load_sent_id_workbook(path: Path) -> dict[str, pd.DataFrame]:
    workbook = pd.ExcelFile(make_readable_copy(path))
    sheets: dict[str, pd.DataFrame] = {}
    for sheet_name in workbook.sheet_names:
        df = pd.read_excel(workbook, sheet_name=sheet_name)
        sheets[sheet_name] = df
    return sheets


def build_training_token_sequences(train_df: pd.DataFrame) -> list[list[str]]:
    sequences: list[list[str]] = []
    for row in train_df.itertuples(index=False):
        script = normalize_script_text(getattr(row, "SCRIPT_CONTIN", ""))
        if not script:
            continue
        raw_state4 = getattr(row, "STATE_4", "")
        tokens = tokens_from_state4(script, raw_state4)
        if tokens:
            sequences.append(tokens)
    return sequences


def format_display_sentence(tokens: list[str] | tuple[str, ...]) -> str:
    return " ".join(token.upper() for token in tokens)


def make_excel_sheet_name(base_name: str, chunk_index: int = 1) -> str:
    if chunk_index <= 1:
        return base_name[:31]
    suffix = f"_{chunk_index}"
    return f"{base_name[:31 - len(suffix)]}{suffix}"


def run_cm_sentence_experiment(
    sentence_file: Path,
    output_file: Path,
    validator: PhonotacticWordlikenessValidator,
    x_values: list[int],
    max_word_len: int,
    beam_cap: int,
) -> None:
    sheets = load_sent_id_workbook(sentence_file)
    if "train" not in sheets:
        raise SystemExit("The CM sentence experiment expects a 'train' sheet in the workbook.")

    context_model = SentenceContextModel(validator)
    train_sequences = build_training_token_sequences(sheets["train"])
    context_model.fit(train_sequences)

    all_rows: list[dict[str, object]] = []

    for sheet_name, df in sheets.items():
        for row_index, row in enumerate(df.itertuples(index=False), start=1):
            script = normalize_script_text(getattr(row, "SCRIPT_CONTIN", ""))
            if not script:
                continue

            reference_tokens = tokens_from_state4(script, getattr(row, "STATE_4", ""))
            reference_sentence = " ".join(reference_tokens)
            candidates = build_candidate_arcs(script, context_model, max_word_len=max_word_len)
            max_candidate_count = max((len(arcs) for arcs in candidates.values()), default=1)
            effective_k_by_x = {
                x_percent: max(1, min(max_candidate_count, math.ceil(len(script) * x_percent / 100.0)))
                for x_percent in x_values
            }
            unique_results: dict[int, tuple[dict[str, float | int | str], int]] = {}

            for x_percent in x_values:
                top_k = effective_k_by_x[x_percent]
                if top_k not in unique_results:
                    unique_results[top_k] = search_best_sentence(
                        script=script,
                        candidates=candidates,
                        context_model=context_model,
                        top_k=top_k,
                        beam_cap=beam_cap,
                    )
                best_metrics, final_path_count = unique_results[top_k]
                all_rows.append(
                    {
                        "sheet_name": sheet_name,
                        "row_number": row_index,
                        "ART_ID": getattr(row, "ART_ID", ""),
                        "PARA_ID": getattr(row, "PARA_ID", ""),
                        "SENT_ID": getattr(row, "SENT_ID", ""),
                        "input_sentence": str(getattr(row, "SENT_ORI", "")),
                        "input_script_contin": script,
                        "reference_state4_sentence": reference_sentence,
                        "x_percent": x_percent,
                        "top_k_per_start": top_k,
                        "candidate_paths_scored": final_path_count,
                        "output_sentence": best_metrics["output_sentence"],
                        "cm_score": float(best_metrics["cm_score"]),
                        "word_validity": float(best_metrics["word_validity"]),
                        "bigram_validity": float(best_metrics["bigram_validity"]),
                        "character_smoothness": float(best_metrics["character_smoothness"]),
                        "sentence_length_sanity": float(best_metrics["sentence_length_sanity"]),
                        "mean_ortho_score": float(best_metrics["mean_ortho_score"]),
                        "token_count": int(best_metrics["token_count"]),
                    }
                )

            if row_index % 100 == 0:
                print(f"processed {sheet_name}: {row_index} sentences")

    results_df = pd.DataFrame(all_rows)
    if results_df.empty:
        raise SystemExit("No sentence rows were exported. Please check the workbook columns.")

    best_df = (
        results_df.sort_values(
            by=["sheet_name", "row_number", "cm_score", "x_percent"],
            ascending=[True, True, False, True],
        )
        .groupby(["sheet_name", "row_number"], as_index=False)
        .first()
    )

    summary_df = pd.DataFrame(
        [
            {
                "train_sentence_count": len(train_sequences),
                "avg_token_length_mean": context_model.avg_token_length_mean,
                "avg_token_length_std": context_model.avg_token_length_std,
                "bigram_raw_mean": context_model.bigram_raw_mean,
                "bigram_raw_std": context_model.bigram_raw_std,
                "x_values": ",".join(str(value) for value in x_values),
                "max_word_len": max_word_len,
                "beam_cap": beam_cap,
            }
        ]
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="cm_x_results", index=False)
        best_df.to_excel(writer, sheet_name="cm_best_per_sentence", index=False)
        summary_df.to_excel(writer, sheet_name="cm_run_summary", index=False)

    print(f"written: {output_file}")
    print(f"rows: {len(results_df)}")
    print(f"best_rows: {len(best_df)}")


def run_coh_sentence_experiment(
    sentence_file: Path,
    output_file: Path,
    validator: PhonotacticWordlikenessValidator,
    x_values: list[int],
    max_word_len: int,
    beam_cap: int,
    eval_sheets: list[str],
) -> None:
    sheets = load_sent_id_workbook(sentence_file)
    if "train" not in sheets:
        raise SystemExit("The COH experiment expects a 'train' sheet.")

    context_model = SentenceContextModel(validator)
    train_sequences = build_training_token_sequences(sheets["train"])
    context_model.fit(train_sequences)

    workbook = Workbook(write_only=True)
    max_excel_rows = 1_000_000

    best_headers = [
        "sheet_name",
        "row_number",
        "ART_ID",
        "PARA_ID",
        "SENT_ID",
        "inp_sentence",
        "ground_truth",
        "x_percent",
        "top_k_per_start",
        "candidate_count",
        "output",
        "best_coh",
    ]
    summary_headers = [
        "sheet_name",
        "row_number",
        "SENT_ID",
        "x_percent",
        "script_length",
        "top_k_per_start",
        "candidate_count",
    ]
    detail_headers = [
        "sheet_name",
        "row_number",
        "ART_ID",
        "PARA_ID",
        "SENT_ID",
        "inp_sentence",
        "ground_truth",
        "x_percent",
        "possible_sentences",
        "coh",
        "output",
    ]

    best_ws = workbook.create_sheet(title="coh best output")
    best_ws.append(best_headers)
    best_row_total = 0
    detail_row_total = 0
    detail_sheets = {x_percent: workbook.create_sheet(title=f"x={x_percent}") for x_percent in x_values}
    for worksheet in detail_sheets.values():
        worksheet.append(detail_headers)

    target_sheet_names = [sheet_name for sheet_name in eval_sheets if sheet_name in sheets]
    if not target_sheet_names:
        raise SystemExit("None of the requested COH evaluation sheets were found in the workbook.")

    for sheet_name in target_sheet_names:
        df = sheets[sheet_name]
        for row_index, row in enumerate(df.itertuples(index=False), start=1):
            script = normalize_script_text(getattr(row, "SCRIPT_CONTIN", ""))
            if not script:
                continue

            sent_ori = str(getattr(row, "SENT_ORI", ""))
            reference_tokens = tokens_from_state4(script, getattr(row, "STATE_4", ""))
            reference_sentence = format_display_sentence(reference_tokens) if reference_tokens else sent_ori.upper()
            input_sentence = sent_ori.upper() if sent_ori else script.upper()
            candidates_arcs = build_candidate_arcs(script, context_model, max_word_len=max_word_len)
            max_candidate_count = max((len(arcs) for arcs in candidates_arcs.values()), default=1)
            cached_candidates: dict[int, list[tuple[str, ...]]] = {}

            for x_percent in x_values:
                top_k = max(1, min(max_candidate_count, math.ceil(len(script) * x_percent / 100.0)))

                if top_k not in cached_candidates:
                    cached_candidates[top_k] = search_all_candidates(
                        script=script,
                        candidates=candidates_arcs,
                        context_model=context_model,
                        top_k=top_k,
                        beam_cap=beam_cap,
                    )

                candidate_token_sequences = cached_candidates[top_k]
                if not candidate_token_sequences:
                    candidate_token_sequences = [(script,)]

                scored_candidates = []
                for tokens in candidate_token_sequences:
                    coh = context_model.compute_coh_score(tokens)
                    scored_candidates.append(
                        {
                            "possible_sentence": format_display_sentence(tokens if isinstance(tokens, tuple) else tuple(tokens)),
                            "coh_score": coh,
                            "token_count": len(tokens),
                        }
                    )

                scored_candidates.sort(
                    key=lambda item: (-item["coh_score"], item["token_count"], item["possible_sentence"])
                )

                best_candidate = scored_candidates[0]
                best_ws.append(
                    [
                        sheet_name,
                        row_index,
                        getattr(row, "ART_ID", ""),
                        getattr(row, "PARA_ID", ""),
                        getattr(row, "SENT_ID", ""),
                        input_sentence,
                        reference_sentence,
                        x_percent,
                        top_k,
                        len(scored_candidates),
                        best_candidate["possible_sentence"],
                        best_candidate["coh_score"],
                    ]
                )
                best_row_total += 1

                detail_ws = detail_sheets[x_percent]
                for i, cand in enumerate(scored_candidates):
                    detail_ws.append(
                        [
                            sheet_name if i == 0 else "",
                            row_index if i == 0 else "",
                            getattr(row, "ART_ID", "") if i == 0 else "",
                            getattr(row, "PARA_ID", "") if i == 0 else "",
                            getattr(row, "SENT_ID", "") if i == 0 else "",
                            input_sentence if i == 0 else "",
                            reference_sentence if i == 0 else "",
                            x_percent if i == 0 else "",
                            cand["possible_sentence"],
                            cand["coh_score"],
                            best_candidate["possible_sentence"] if i == 0 else "",
                        ],
                    )
                    detail_row_total += 1

            if row_index % 50 == 0:
                print(f"processed {sheet_name}: {row_index} sentences")

    if best_row_total == 0:
        print("Warning: No results generated for COH experiment.")
        return

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    print(f"written Excel: {output_file}")
    print(f"best_rows: {best_row_total}")
    print(f"detail_rows: {detail_row_total}")


def parse_x_values(raw_value: str) -> list[int]:
    values: list[int] = []
    for part in raw_value.split(","):
        part = part.strip()
        if not part:
            continue
        value = int(part)
        if value < 1 or value > 100:
            raise SystemExit("CM x values must be between 1 and 100.")
        values.append(value)
    if not values:
        raise SystemExit("Please provide at least one CM x value.")
    return values


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Phonotactic and wordlikeness validator.")
    parser.add_argument("--train", type=Path, default=Path("datasets/train_word_count.xlsx"))
    parser.add_argument("--word", type=str, help="Optional word to score.")
    parser.add_argument("--eval-file", type=Path, help="Optional Excel file to evaluate.")
    parser.add_argument("--score-only", action="store_true")
    parser.add_argument("--cm-sentence-file", type=Path, help="Run the CM-only sentence experiment on a SENT_ID workbook.")
    parser.add_argument("--cm-output-file", type=Path, help="Output Excel path for the CM-only sentence experiment.")
    parser.add_argument("--cm-x-values", type=str, default="10,20,30,40,50,60,70,80,90,100")
    parser.add_argument("--cm-max-word-len", type=int, default=20)
    parser.add_argument("--cm-beam-cap", type=int, default=160)
    parser.add_argument("--coh-sentence-file", type=Path, help="Run the COH-only sentence experiment.")
    parser.add_argument("--coh-output-file", type=Path, help="Output Excel path for the COH experiment.")
    parser.add_argument("--coh-eval-sheets", type=str, default="test")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    train_words = load_words(args.train)
    validator = PhonotacticWordlikenessValidator()
    validator.fit(train_words)

    if args.coh_sentence_file is not None:
        output_file = args.coh_output_file
        if output_file is None:
            output_file = Path("results/spreadsheets/coh_sentence_results.xlsx")
        run_coh_sentence_experiment(
            sentence_file=args.coh_sentence_file,
            output_file=output_file,
            validator=validator,
            x_values=parse_x_values(args.cm_x_values),
            max_word_len=args.cm_max_word_len,
            beam_cap=args.cm_beam_cap,
            eval_sheets=[sheet.strip() for sheet in args.coh_eval_sheets.split(",") if sheet.strip()],
        )
        return

    if args.cm_sentence_file is not None:
        output_file = args.cm_output_file
        if output_file is None:
            output_file = Path("results/spreadsheets/cm_sentence_x_results.xlsx")
        run_cm_sentence_experiment(
            sentence_file=args.cm_sentence_file,
            output_file=output_file,
            validator=validator,
            x_values=parse_x_values(args.cm_x_values),
            max_word_len=args.cm_max_word_len,
            beam_cap=args.cm_beam_cap,
        )
        return

    if args.word is not None:
        result = validator.score_word(args.word)
        if args.score_only:
            print(f"{result['score']:.6f}")
            return

        print(f"training_words: {len(train_words)}")
        print(f"word: {result['word']}")
        print(f"is_wordlike: {result['is_wordlike']}")
        print(f"score: {result['score']:.6f}")
        print(f"ngram_score: {result['ngram_score']:.6f}")
        print(f"chunk_score: {result['chunk_score']:.6f}")
        print(f"shape_score: {result['shape_score']:.6f}")
        print(f"familiarity_score: {result['familiarity_score']:.6f}")
        return

    if args.eval_file is not None:
        metrics = evaluate_dataset(args.eval_file, validator)
        print(f"training_words: {len(train_words)}")
        print(f"eval_file: {args.eval_file}")
        print(f"rows: {metrics['rows']}")
        print(f"total_weight: {metrics['total_weight']}")
        print(f"accepted_row_fraction: {metrics['accepted_row_fraction']:.6f}")
        print(f"accepted_weight_fraction: {metrics['accepted_weight_fraction']:.6f}")
        print(f"average_score: {metrics['average_score']:.6f}")
        print(f"weighted_average_score: {metrics['weighted_average_score']:.6f}")


if __name__ == "__main__":
    main()
