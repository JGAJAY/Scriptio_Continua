from __future__ import annotations

from pathlib import Path

import openpyxl

from character_backoff_language_model import (
    SimpleBackoffChainModel,
    load_words,
    normalize_word,
)
from character_backoff_pronounceability_validator import CharacterPronounceabilityValidator
from hybrid_lexical_character_pronounceability_validator import (
    HybridWordValidator as CharacterHybridWordValidator,
)
from hybrid_lexical_pronounceability_validator import HybridWordValidator as SyllableHybridWordValidator
from lexical_frequency_validator import build_lexicon, validate_word
from orthographic_phonotactic_validator import PhonotacticWordlikenessValidator
from phoneme_sequence_wordlikeness_validator import PhonemeWordlikenessValidator
from syllable_backoff_pronounceability_validator import PronounceabilityValidator


ROOT = Path(__file__).resolve().parent
DATASET_PATH = ROOT / "test dataset.xlsx"
TRAIN_PATH = ROOT / "datasets" / "train_word_count.xlsx"

SCRIPT_COLUMNS = [
    "character_backoff_language_model.py",
    "character_backoff_pronounceability_validator.py",
    "export_hybrid_validation_results.py",
    "hybrid_lexical_character_pronounceability_validator.py",
    "hybrid_lexical_pronounceability_validator.py",
    "lexical_frequency_validator.py",
    "orthographic_phonotactic_validator.py",
    "phoneme_sequence_wordlikeness_validator.py",
    "plot_score_distributions.py",
    "syllable_backoff_pronounceability_validator.py",
]

UTILITY_VALUE = "utility_script_no_per_word_output"


def build_scorers() -> dict[str, callable]:
    train_words = load_words(TRAIN_PATH)
    lexicon = build_lexicon(wordlist="large")

    character_model = SimpleBackoffChainModel()
    character_model.fit(train_words)

    character_pronounceability = CharacterPronounceabilityValidator()
    character_pronounceability.fit(train_words)

    syllable_pronounceability = PronounceabilityValidator()
    syllable_pronounceability.fit(train_words)

    character_hybrid = CharacterHybridWordValidator(
        lexicon=lexicon,
        pronounceability=character_pronounceability,
    )
    syllable_hybrid = SyllableHybridWordValidator(
        lexicon=lexicon,
        pronounceability=syllable_pronounceability,
    )

    orthographic = PhonotacticWordlikenessValidator()
    orthographic.fit(train_words)

    phoneme = PhonemeWordlikenessValidator()
    phoneme.fit(train_words)

    def character_backoff_probability(raw_word: str) -> float:
        word = normalize_word(raw_word)
        if not word:
            return 0.0
        probability, _ = character_model.word_probability(word)
        return probability

    return {
        "character_backoff_language_model.py": character_backoff_probability,
        "character_backoff_pronounceability_validator.py": lambda word: character_pronounceability.score_word(word)["score"],
        "export_hybrid_validation_results.py": lambda word: UTILITY_VALUE,
        "hybrid_lexical_character_pronounceability_validator.py": lambda word: character_hybrid.score_word(word)["score"],
        "hybrid_lexical_pronounceability_validator.py": lambda word: syllable_hybrid.score_word(word)["score"],
        "lexical_frequency_validator.py": lambda word: validate_word(word, lexicon)["normalized_score"],
        "orthographic_phonotactic_validator.py": lambda word: orthographic.score_word(word)["score"],
        "phoneme_sequence_wordlikeness_validator.py": lambda word: phoneme.score_word(word)["score"],
        "plot_score_distributions.py": lambda word: UTILITY_VALUE,
        "syllable_backoff_pronounceability_validator.py": lambda word: syllable_pronounceability.score_word(word)["score"],
    }


def get_or_create_column_map(ws) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=column).value
        if header is not None:
            header_map[str(header)] = column

    next_column = ws.max_column + 1
    for script_name in SCRIPT_COLUMNS:
        if script_name not in header_map:
            ws.cell(row=1, column=next_column, value=script_name)
            header_map[script_name] = next_column
            next_column += 1

    return header_map


def main() -> None:
    workbook = openpyxl.load_workbook(DATASET_PATH)
    worksheet = workbook[workbook.sheetnames[0]]
    scorers = build_scorers()
    header_map = get_or_create_column_map(worksheet)

    for row in range(2, worksheet.max_row + 1):
        word = worksheet.cell(row=row, column=1).value
        for script_name in SCRIPT_COLUMNS:
            value = scorers[script_name](word)
            worksheet.cell(row=row, column=header_map[script_name], value=value)

    workbook.save(DATASET_PATH)
    print(f"updated_file: {DATASET_PATH}")
    print(f"rows_processed: {max(0, worksheet.max_row - 1)}")
    print(f"columns_written: {len(SCRIPT_COLUMNS)}")


if __name__ == "__main__":
    main()
