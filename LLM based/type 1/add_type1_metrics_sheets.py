from __future__ import annotations

import math
import re
from collections import Counter
from pathlib import Path
from typing import Iterable

import pandas as pd
from bert_score import score as bert_score
from nltk.translate.bleu_score import SmoothingFunction, sentence_bleu
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
FILES = [
    ROOT / "type1_deepseek-r1_test_results.xlsx",
    ROOT / "type1_gemma4_e2b_test_results.xlsx",
    ROOT / "type1_qwen3.5_4b_test_results.xlsx",
]

TOKEN_RE = re.compile(r"[a-z0-9]+")
LABELS = ["B", "M", "E", "S"]
RESULT_SHEET = "results"
METRICS_SHEET = "metrics"
AVG_SHEET = "avg sheet"
OUTPUT_SUFFIX = "_with_metrics"


def tokenize_text(value: object) -> list[str]:
    return TOKEN_RE.findall(str(value or "").lower())


def parse_state4(value: object, n_chars: int) -> list[str]:
    tags = re.findall(r"[SBEMI]", str(value or "").upper())
    tags = ["M" if tag == "I" else tag for tag in tags]
    if len(tags) < n_chars:
        tags += ["M"] * (n_chars - len(tags))
    return tags[:n_chars]


def lcs_length(a: list[str], b: list[str]) -> int:
    if not a or not b:
        return 0
    prev = [0] * (len(b) + 1)
    for token_a in a:
        curr = [0]
        for j, token_b in enumerate(b, start=1):
            if token_a == token_b:
                curr.append(prev[j - 1] + 1)
            else:
                curr.append(max(curr[-1], prev[j]))
        prev = curr
    return prev[-1]


def rouge_n_f1(reference: list[str], candidate: list[str], n: int) -> float:
    if len(reference) < n or len(candidate) < n:
        return 0.0
    ref_counts = Counter(tuple(reference[i : i + n]) for i in range(len(reference) - n + 1))
    cand_counts = Counter(tuple(candidate[i : i + n]) for i in range(len(candidate) - n + 1))
    overlap = sum(min(count, cand_counts[gram]) for gram, count in ref_counts.items())
    ref_total = sum(ref_counts.values())
    cand_total = sum(cand_counts.values())
    if ref_total == 0 or cand_total == 0 or overlap == 0:
        return 0.0
    precision = overlap / cand_total
    recall = overlap / ref_total
    return 2 * precision * recall / (precision + recall)


def rouge_l_f1(reference: list[str], candidate: list[str]) -> float:
    if not reference or not candidate:
        return 0.0
    lcs = lcs_length(reference, candidate)
    if lcs == 0:
        return 0.0
    precision = lcs / len(candidate)
    recall = lcs / len(reference)
    return 2 * precision * recall / (precision + recall)


def meteor_score_simple(reference: list[str], candidate: list[str]) -> float:
    if not reference or not candidate:
        return 0.0

    ref_positions: dict[str, list[int]] = {}
    for index, token in enumerate(reference):
        ref_positions.setdefault(token, []).append(index)

    matched_ref_indices: list[int] = []
    used_counter: Counter[str] = Counter()
    matches = 0
    for token in candidate:
        positions = ref_positions.get(token, [])
        use_index = used_counter[token]
        if use_index < len(positions):
            matched_ref_indices.append(positions[use_index])
            used_counter[token] += 1
            matches += 1

    if matches == 0:
        return 0.0

    precision = matches / len(candidate)
    recall = matches / len(reference)
    denom = recall + 9 * precision
    f_mean = (10 * precision * recall / denom) if denom else 0.0

    chunks = 1
    for left, right in zip(matched_ref_indices, matched_ref_indices[1:]):
        if right != left + 1:
            chunks += 1
    penalty = 0.5 * ((chunks / matches) ** 3)
    return (1.0 - penalty) * f_mean


def macro_state_metrics(reference: list[str], predicted: list[str]) -> dict[str, float]:
    max_len = max(len(reference), len(predicted), 1)
    if len(reference) < max_len:
        reference = reference + ["M"] * (max_len - len(reference))
    if len(predicted) < max_len:
        predicted = predicted + ["M"] * (max_len - len(predicted))

    accuracy = sum(int(left == right) for left, right in zip(reference, predicted)) / max_len

    precisions: list[float] = []
    recalls: list[float] = []
    f1s: list[float] = []
    for label in LABELS:
        tp = sum(1 for ref, pred in zip(reference, predicted) if ref == label and pred == label)
        fp = sum(1 for ref, pred in zip(reference, predicted) if ref != label and pred == label)
        fn = sum(1 for ref, pred in zip(reference, predicted) if ref == label and pred != label)
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0
        precisions.append(precision)
        recalls.append(recall)
        f1s.append(f1)

    return {
        "f1": sum(f1s) / len(f1s),
        "precision": sum(precisions) / len(precisions),
        "accuracy": accuracy,
        "recall": sum(recalls) / len(recalls),
    }


def sentence_metrics(reference_sentence: str, output_sentence: str) -> dict[str, float]:
    ref_tokens = tokenize_text(reference_sentence)
    out_tokens = tokenize_text(output_sentence)
    if ref_tokens and out_tokens:
        bleu = sentence_bleu([ref_tokens], out_tokens, smoothing_function=SmoothingFunction().method1)
    else:
        bleu = 0.0
    return {
        "bleu": bleu,
        "rouge_1_f1": rouge_n_f1(ref_tokens, out_tokens, 1),
        "rouge_2_f1": rouge_n_f1(ref_tokens, out_tokens, 2),
        "rouge_l_f1": rouge_l_f1(ref_tokens, out_tokens),
        "meteor": meteor_score_simple(ref_tokens, out_tokens),
    }


def compute_metric_rows(df: pd.DataFrame) -> pd.DataFrame:
    records = df.to_dict(orient="records")
    bert_inputs: list[tuple[str, str]] = []

    for record in records:
        reference_sentence = str(record.get("ground truth sent") or "")
        output_sentence = str(record.get("output direct output") or "")
        script = str(record.get("scriptio continua sentence") or "")
        ground_truth_state = parse_state4(record.get("ground truth 4 state"), len(script))
        output_state = parse_state4(record.get("output 4 state"), len(script))

        text_scores = sentence_metrics(reference_sentence, output_sentence)
        state_scores = macro_state_metrics(ground_truth_state, output_state)

        record.update(text_scores)
        record.update(state_scores)
        bert_inputs.append((reference_sentence, output_sentence))

    if bert_inputs:
        candidates = [pair[1] for pair in bert_inputs]
        references = [pair[0] for pair in bert_inputs]
        precision, recall, f1 = bert_score(
            candidates,
            references,
            lang="en",
            model_type="distilbert-base-uncased",
            verbose=False,
            device="cpu",
        )
        for record, p, r, f in zip(records, precision.tolist(), recall.tolist(), f1.tolist()):
            record["bertscore_p"] = float(p)
            record["bertscore_r"] = float(r)
            record["bertscore_f1"] = float(f)

    ordered_columns = list(df.columns) + [
        "bleu",
        "rouge_1_f1",
        "rouge_2_f1",
        "rouge_l_f1",
        "meteor",
        "bertscore_p",
        "bertscore_r",
        "bertscore_f1",
        "f1",
        "precision",
        "accuracy",
        "recall",
    ]
    return pd.DataFrame(records, columns=ordered_columns)


def build_avg_sheet_df(metrics_df: pd.DataFrame) -> pd.DataFrame:
    metric_columns = [
        "bleu",
        "rouge_1_f1",
        "rouge_2_f1",
        "rouge_l_f1",
        "meteor",
        "bertscore_p",
        "bertscore_r",
        "bertscore_f1",
        "f1",
        "precision",
        "accuracy",
        "recall",
    ]
    averages = {"source_sheet": METRICS_SHEET}
    for column in metric_columns:
        averages[column] = float(metrics_df[column].mean()) if column in metrics_df else math.nan
    return pd.DataFrame([averages], columns=["source_sheet", *metric_columns])


def remove_sheet_if_present(workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]


def iter_target_files() -> Iterable[Path]:
    for path in FILES:
        if path.exists():
            yield path


def process_workbook(path: Path) -> Path:
    results_df = pd.read_excel(path, sheet_name=RESULT_SHEET)
    metrics_df = compute_metric_rows(results_df)
    avg_df = build_avg_sheet_df(metrics_df)

    output_path = path.with_name(f"{path.stem}{OUTPUT_SUFFIX}{path.suffix}")
    workbook = load_workbook(path)
    remove_sheet_if_present(workbook, METRICS_SHEET)
    remove_sheet_if_present(workbook, AVG_SHEET)
    workbook.save(output_path)
    workbook.close()

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        metrics_df.to_excel(writer, sheet_name=METRICS_SHEET, index=False)
        avg_df.to_excel(writer, sheet_name=AVG_SHEET, index=False)

    return output_path


def main() -> None:
    for path in iter_target_files():
        output_path = process_workbook(path)
        print(f"Created {output_path}")


if __name__ == "__main__":
    main()
